import sys
import os
import subprocess
import openpyxl
import colorama as cor
import socket
import threading

cor.init()

# Verifica se o número correto de argumentos foi fornecido
if len(sys.argv) < 2:
    print('Uso: python script.py <caminho_do_excel>')
    sys.exit()

file_path = sys.argv[1]  # Caminho do arquivo Excel

if not os.path.exists(file_path):
    print(f'O arquivo {file_path} não existe.')
    sys.exit()

# Portas específicas a serem verificadas
PORTAS_ESPECIFICAS = [22, 80, 443, 3040, 5001]

# Função para obter o registro PTR (DNS reverso) de um IP
def get_ptr_record(ip):
    try:
        return socket.gethostbyaddr(ip)[0]
    except (socket.herror, socket.gaierror):
        return "N/A"
    except Exception as e:
        return f"Erro: {str(e)}"

# Função para verificar se uma porta está aberta
def check_port(ip, port):
    try:
        client = socket.socket()
        client.settimeout(0.05)
        client.connect((ip, port))
        client.close()
        return True
    except:
        return False

# Função para verificar as portas específicas
def check_ports(ip, portas_abertas, lock):
    for port in PORTAS_ESPECIFICAS:
        if check_port(ip, port):
            with lock:
                portas_abertas.append(port)

# Função para verificar se um host está ativo e suas portas abertas
def check_host(entrada, ativos, desativos, lock):
    try:
        # Resolução DNS
        ip = socket.gethostbyname(entrada)
        
        # Resolução DNS reversa (PTR)
        dns_name = get_ptr_record(ip)

    except (socket.gaierror, Exception) as e:
        ip = f"Erro: {str(e)}"
        dns_name = "N/A"

    try:
        # Executa o comando ping
        ping = subprocess.Popen(
            ["C:\\Windows\\System32\\ping.exe", "-n", "1", entrada],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            shell=True
        )
        stdout, stderr = ping.communicate()
        
        if ping.returncode == 0:
            portas_abertas = []
            port_threads = []

            # Cria uma thread para verificar as portas específicas
            thread = threading.Thread(
                target=check_ports,
                args=(ip, portas_abertas, lock)
            )
            port_threads.append(thread)
            thread.start()

            # Aguarda a thread de verificação de portas terminar
            for thread in port_threads:
                thread.join()

            # Adiciona o host à lista de ativos
            with lock:
                ativos.append((entrada, ip, dns_name, portas_abertas))
        else:
            # Adiciona o host à lista de inativos
            with lock:
                desativos.append((entrada, ip, dns_name))

    except Exception as e:
        print(f'Erro ao tentar fazer ping em {entrada}: {e}')

try:
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    ativos = []
    desativos = []
    lock = threading.Lock()  # Lock para sincronizar o acesso às listas

    threads = []
    for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
        for cell in row:
            entrada = cell.value
            if entrada:
                # Cria uma thread para cada host
                thread = threading.Thread(
                    target=check_host,
                    args=(entrada, ativos, desativos, lock))
                threads.append(thread)
                thread.start()

    # Aguarda todas as threads terminarem
    for thread in threads:
        thread.join()

    print(f"{cor.Fore.BLUE} \n######################### Resultados Finais ###########################{cor.Fore.RESET}")
    
    print("\nATIVOS:\n")
    for entrada, ip, dns, portas in ativos:
        if dns == 'N/A' and ip.startswith('201.46'):
            dns = 'Altatech'
        elif dns.startswith('srv'):
            dns = 'Hostinger'
        elif dns.startswith('vmi'):
            dns = 'Contabo'
        else:
            dns = 'DNS não reconhecido.'
        print(f'{cor.Fore.GREEN}[+]{cor.Fore.RESET} {entrada}')
        print(f'{cor.Fore.GREEN}[+]{cor.Fore.RESET} IP: {ip}')
        print(f'{cor.Fore.GREEN}[+]{cor.Fore.RESET} DNS: {dns}')
        print(f'{cor.Fore.GREEN}[+]{cor.Fore.RESET} Portas abertas: {portas}\n')

    print("NÃO ATIVOS:\n")
    for entrada, ip, dns in desativos:
        print(f'{cor.Fore.RED}[-]{cor.Fore.RESET} {entrada}\n')

except Exception as e:
    print(f'Erro: {e}')
    sys.exit()
